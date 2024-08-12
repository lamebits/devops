from sys import path
import pandas as pd
import openpyxl
from fpdf import FPDF
from pathlib import Path

root_dir = Path('Outputfiles')
for path in root_dir.rglob("*.pdf"):
  if path.exists():
    path.unlink()
    print(f"File '{path.name}' has been deleted.")
  else:
    print(f"File '{path.name}' does not exist.")

df = pd.read_excel("Inputfiles/data.xlsx")

for index, row in df.iterrows():
  pdf = FPDF(orientation="P", unit="pt", format="A4")
  pdf.add_page()
  pdf.set_font(family="Times", style="B", size=24)
  pdf.cell(w=0, h=15, txt=row['name'], align='c', ln=1)

  for col in df.columns[1:]:
    pdf.set_font(family="Times", style="B", size=14)
    pdf.cell(w=100, h=25, txt=f'{col.title()}', border=1)
    pdf.set_font(family="Times", style="B", size=14)
    pdf.cell(w=100, h=25, txt=f'{row[col]}', border=1, ln=1)

  pdf.output("Outputfiles/" + f"{row['name']}.pdf")
