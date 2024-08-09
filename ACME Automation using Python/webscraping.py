from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pandas as pd
import os
from openpyxl import load_workbook
from constants import ACME_URL,IMPLICIT_DELAY,EMAIL_XPATH,PASSWORD_XPATH,SUBMIT_XPATH,WORKITEM_TABLE_XPATH,WORKITEM_XPATH,INPUT_FILE_PATH,AUTHENTICATION,ERROR_MESSAGE
from dotenv import load_dotenv
load_dotenv()

def getDriver():
    # configuration of chrome browser
    options = webdriver.ChromeOptions()
    options.add_argument("disable-infobars")
    options.add_argument("start-maximized")
    options.add_argument("disable-dev-shm-usage")
    options.add_argument("no-sandbox")
    options.add_argument("disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])

    try:
        driver = webdriver.Chrome(options=options)
        # Open URL
        driver.get(ACME_URL)
        return driver
    except Exception as e:
        print(f"{ERROR_MESSAGE.get("DRIVER_INITIALIZATION")}{e}")
        return None

def acmeLogin(driver):
    emailId = os.getenv("ACME_EMAILID")
    passwordID = os.getenv("ACME_PASSWORD")
    if driver is None:
        print(ERROR_MESSAGE.get("DRIVER_INITIALIZATION"))
        return

    wait = WebDriverWait(driver, IMPLICIT_DELAY)
    # Login ACME Portal using Credentials
    try:
        acmeHomePage = wait.until(EC.presence_of_element_located((By.XPATH, EMAIL_XPATH)))
        print("ACME Home Page Found")
    except TimeoutException:
        #print("ACME Home Page Not Found")
        print(ERROR_MESSAGE["ACME_LOGIN"])
        driver.quit()
        return
    
    email = driver.find_element(by="xpath",value=EMAIL_XPATH)
    email.send_keys(emailId)
    
    password = driver.find_element(by="xpath",value=PASSWORD_XPATH)
    password.send_keys(passwordID)
    
    submitbtn = driver.find_element(by="xpath",value=SUBMIT_XPATH)
    submitbtn.click()
    
    try:
        authenticateLogin = wait.until(EC.presence_of_element_located((By.XPATH, AUTHENTICATION(emailId))))
    except Exception as e:
        print(f"Login Unsuccessfully \n {e}")
        driver.quit()
    
    autherEmailID = driver.find_element(by="xpath",value=AUTHENTICATION(emailId))
    print(autherEmailID.text)
    if autherEmailID.text == emailId:
        print("ACME Portal Login Successfully")
    else:
        print("ACME Portal Login Unsuccessfully")

def fetchTable(driver):  
    #Click on Worrk Item Tab
    workItems = driver.find_element(by="xpath",value=WORKITEM_XPATH)
    workItems.click()

    #Fetch Table Data
    table = driver.find_element(by="xpath",value=WORKITEM_TABLE_XPATH)
    rows = table.find_elements(By.TAG_NAME, 'tr')

    #Declare a table data list
    table_data = []
    # For Table Header Data
    cells = table.find_elements(By.TAG_NAME, 'th')  # Use 'th' if it's a header row
    row_data = [cell.text for cell in cells]
    print(row_data)
    table_data.append(row_data)

    #For Table Row Data
    for row2 in rows:
        cells2 = row2.find_elements(By.TAG_NAME, 'td')  # Use 'thd' if it's row data
        row_data2 = [cell2.text for cell2 in cells2]
        print(row_data2)
        table_data.append(row_data2)

    #Convert Panda Data to dataFrame
    df = pd.DataFrame(table_data)
    print(df)
    file_path = INPUT_FILE_PATH
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        # If Excel already exists then append the data
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, header=False, index=False)
    else:
        #Create new excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)

    print(f"Data successfully written to {file_path}")

def webscraping_main():
    driver = getDriver()
    
    try:
        acmeLogin(driver)
    except Exception as e:
        print(f"ACME Portal Login Unsuccessfully \n {e}")
        driver.quit()
    
    try:
        fetchTable(driver)
    except Exception as e:
        print(f"Failed to fetch data and prepare excel {e}")
        driver.quit()
    
    driver.quit()

if __name__ == "__main__":
    webscraping_main()